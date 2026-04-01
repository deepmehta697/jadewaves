from __future__ import annotations

import argparse
import json
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.workbook.properties import CalcProperties
from openpyxl.worksheet.page import PageMargins


ROOT = Path(__file__).resolve().parent
OUTPUT_DIR = ROOT / "output" / "xlsx"

THIN = Side(style="thin", color="000000")
MEDIUM = Side(style="medium", color="000000")
SOFT = Side(style="thin", color="D8D6D0")
DARK = Side(style="thin", color="10263A")
INVISIBLE = Side(style=None, color="FFFFFF")

BLUE_FILL = PatternFill("solid", "D9E2F3")
SECTION_FILL = PatternFill("solid", "D9E2F3")
EDIT_FILL = PatternFill("solid", "FFF2CC")
FORMULA_FILL = PatternFill("solid", "F2F2F2")
WHITE_FILL = PatternFill("solid", "FFFFFF")
NAVY_FILL = PatternFill("solid", "10263A")
PANEL_FILL = PatternFill("solid", "FCFBF8")
PANEL_ALT_FILL = PatternFill("solid", "F7F4EE")
ACCENT_FILL = PatternFill("solid", "E9F3F1")
GOLD_FILL = PatternFill("solid", "F5E8C8")
PREMIUM_EDIT_FILL = PatternFill("solid", "FFF7E3")
PREMIUM_FORMULA_FILL = PatternFill("solid", "F2F0EB")

TITLE_FONT = Font(name="Times New Roman", size=18, bold=False)
COMPANY_FONT = Font(name="Calibri", size=11, bold=True)
SMALL_FONT = Font(name="Calibri", size=8)
SMALL_BOLD = Font(name="Calibri", size=8, bold=True)
BODY_FONT = Font(name="Calibri", size=9)
BODY_BOLD = Font(name="Calibri", size=9, bold=True)
TINY_FONT = Font(name="Calibri", size=7)
LINK_FONT = Font(name="Calibri", size=8, color="0563C1", underline="single")
PREMIUM_TITLE_FONT = Font(name="Times New Roman", size=22, bold=True, color="FFFFFF")
PREMIUM_META_FONT = Font(name="Calibri", size=8, bold=True, color="C7A35B")
PREMIUM_MUTED_FONT = Font(name="Calibri", size=8, color="D7E0EA")
PREMIUM_CARD_TITLE_FONT = Font(name="Calibri", size=8, bold=True, color="5A6A7F")
PREMIUM_BODY_FONT = Font(name="Calibri", size=8.5, color="13243C")
PREMIUM_BODY_BOLD = Font(name="Calibri", size=8.5, bold=True, color="13243C")
PREMIUM_SMALL_FONT = Font(name="Calibri", size=7.5, color="5A6A7F")
PREMIUM_TABLE_HEADER_FONT = Font(name="Calibri", size=7.2, bold=True, color="FFFFFF")
PREMIUM_WHITE_FONT = Font(name="Calibri", size=8.5, bold=True, color="FFFFFF")
PREMIUM_TOTAL_FONT = Font(name="Calibri", size=10, bold=True, color="13243C")

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)
CENTER_TOP = Alignment(horizontal="center", vertical="top", wrap_text=True)
RIGHT_CENTER = Alignment(horizontal="right", vertical="center", wrap_text=True)


def text(value: object, default: str = "") -> str:
    if value is None:
        return default
    return str(value).strip()


def lines(value: object) -> list[str]:
    if value is None:
        return []
    if isinstance(value, list):
        return [text(item) for item in value if text(item)]
    line = text(value)
    return [line] if line else []


def to_decimal(value: object, default: str = "0") -> Decimal:
    raw = text(value, default).replace(",", "")
    if not raw:
        raw = default
    try:
        return Decimal(raw)
    except InvalidOperation:
        return Decimal(default)


def amount_to_words(value: Decimal, currency: str) -> str:
    units = [
        "",
        "One",
        "Two",
        "Three",
        "Four",
        "Five",
        "Six",
        "Seven",
        "Eight",
        "Nine",
        "Ten",
        "Eleven",
        "Twelve",
        "Thirteen",
        "Fourteen",
        "Fifteen",
        "Sixteen",
        "Seventeen",
        "Eighteen",
        "Nineteen",
    ]
    tens = ["", "", "Twenty", "Thirty", "Forty", "Fifty", "Sixty", "Seventy", "Eighty", "Ninety"]

    def under_thousand(number: int) -> str:
        parts: list[str] = []
        hundreds = number // 100
        remainder = number % 100
        if hundreds:
            parts.append(f"{units[hundreds]} Hundred")
        if remainder:
            if remainder < 20:
                parts.append(units[remainder])
            else:
                parts.append(tens[remainder // 10])
                if remainder % 10:
                    parts.append(units[remainder % 10])
        return " ".join(part for part in parts if part)

    whole = int(value.quantize(Decimal("1")))
    if whole == 0:
        return f"{currency.upper()} Zero Only"

    parts: list[str] = []
    thousands = whole // 1000
    remainder = whole % 1000
    if thousands:
        parts.append(f"{under_thousand(thousands)} Thousand")
    if remainder:
        parts.append(under_thousand(remainder))

    return f"{currency.upper()} {' '.join(parts)} Only"


def apply_border_range(ws, cell_range: str, fill=None, font=None, alignment=None) -> None:
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell = ws.cell(row=row, column=col)
            left = MEDIUM if col == min_col else THIN
            right = MEDIUM if col == max_col else THIN
            top = MEDIUM if row == min_row else THIN
            bottom = MEDIUM if row == max_row else THIN
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)
            if fill is not None:
                cell.fill = fill
            if font is not None:
                cell.font = font
            if alignment is not None:
                cell.alignment = alignment


def merge_box(ws, cell_range: str, value: object = "", fill=None, font=None, alignment=None) -> None:
    apply_border_range(ws, cell_range, fill=fill, font=font, alignment=alignment)
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    if not (min_col == max_col and min_row == max_row):
        ws.merge_cells(cell_range)
    cell = ws[cell_range.split(":")[0]]
    cell.value = value
    if fill is not None:
        cell.fill = fill
    if font is not None:
        cell.font = font
    if alignment is not None:
        cell.alignment = alignment


def write_cell(ws, ref: str, value: object, fill=None, font=None, alignment=None, number_format: str | None = None) -> None:
    cell = ws[ref]
    cell.value = value
    if fill is not None:
        cell.fill = fill
    if font is not None:
        cell.font = font
    if alignment is not None:
        cell.alignment = alignment
    if number_format is not None:
        cell.number_format = number_format


def configure_sheet(ws, print_area: str, landscape: bool = False) -> None:
    ws.sheet_view.showGridLines = False
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = "landscape" if landscape else "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = True
    ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.35, bottom=0.35, header=0.15, footer=0.15)
    ws.print_area = print_area


def add_image(ws, image_path: Path | None, anchor: str, width: int, height: int) -> None:
    if image_path is None or not image_path.exists():
        return
    image = XLImage(str(image_path))
    image.width = width
    image.height = height
    ws.add_image(image, anchor)


def apply_premium_range(
    ws,
    cell_range: str,
    *,
    fill=None,
    font=None,
    alignment=None,
    outer_side: Side = SOFT,
    inner_side: Side = SOFT,
) -> None:
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell = ws.cell(row=row, column=col)
            left = outer_side if col == min_col else inner_side
            right = outer_side if col == max_col else inner_side
            top = outer_side if row == min_row else inner_side
            bottom = outer_side if row == max_row else inner_side
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)
            if fill is not None:
                cell.fill = fill
            if font is not None:
                cell.font = font
            if alignment is not None:
                cell.alignment = alignment


def merge_premium(
    ws,
    cell_range: str,
    value: object = "",
    *,
    fill=None,
    font=None,
    alignment=None,
    outer_side: Side = SOFT,
    inner_side: Side = SOFT,
) -> None:
    apply_premium_range(
        ws,
        cell_range,
        fill=fill,
        font=font,
        alignment=alignment,
        outer_side=outer_side,
        inner_side=inner_side,
    )
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    if not (min_col == max_col and min_row == max_row):
        ws.merge_cells(cell_range)
    cell = ws[cell_range.split(":")[0]]
    cell.value = value
    if fill is not None:
        cell.fill = fill
    if font is not None:
        cell.font = font
    if alignment is not None:
        cell.alignment = alignment


def short_brand_name(company_name: str) -> str:
    if "Jade Waves" in company_name:
        return "Jade Waves"
    words = company_name.split()
    return " ".join(words[:2]) if words else company_name


def normalize_data(data: dict, data_path: Path) -> dict:
    company = data.get("company", {})
    invoice = data.get("invoice", {})
    consignee = data.get("consignee", {})
    notify = data.get("notify") or consignee
    bank = data.get("bank", {})
    packing = data.get("packing_list", {})
    items = data.get("items", [])

    logo_path = Path(text(company.get("logo_path"), "assets/jade-waves-logo-transparent.png"))
    if not logo_path.is_absolute():
        logo_path = (ROOT / logo_path).resolve()
    signature_path = Path(text(company.get("signature_path"), "assets/coa-silica-sign-stamp.png"))
    if not signature_path.is_absolute():
        signature_path = (ROOT / signature_path).resolve()

    invoice_items = []
    for index, item in enumerate(items, start=1):
        quantity = to_decimal(item.get("quantity"), "0")
        rate = to_decimal(item.get("rate"), "0")
        igst = to_decimal(item.get("igst_rate"), "0")
        description_lines = lines(item.get("description_lines"))
        if not description_lines and text(item.get("description")):
            description_lines.append(text(item.get("description")))
        invoice_items.append(
            {
                "index": index,
                "product": text(item.get("product"), f"Item {index}"),
                "description_lines": description_lines,
                "quantity": quantity,
                "rate": rate,
                "igst_rate": igst,
            }
        )

    amount_words = text(invoice.get("amount_words"))
    if not amount_words:
        total_amount = Decimal("0")
        for item in invoice_items:
            amount = item["quantity"] * item["rate"]
            total_amount += amount * (Decimal("1") + (item["igst_rate"] / Decimal("100")))
        amount_words = amount_to_words(total_amount, text(invoice.get("currency"), "USD"))

    packing_rows = packing.get("rows", [])
    normalized_packing_rows = []
    if packing_rows:
        for row in packing_rows:
            normalized_packing_rows.append(
                {
                    "marks_numbers": text(row.get("marks_numbers")),
                    "packages": text(row.get("packages")),
                    "description_lines": lines(row.get("description_lines")),
                    "quantity_mt": to_decimal(row.get("quantity_mt"), "0"),
                }
            )
    else:
        for item in invoice_items:
            normalized_packing_rows.append(
                {
                    "marks_numbers": text(packing.get("default_marks_numbers")),
                    "packages": text(invoice.get("packages")),
                    "description_lines": [item["product"], *item["description_lines"]],
                    "quantity_mt": item["quantity"],
                }
            )

    return {
        "company": {
            "name": text(company.get("name"), "Jade Waves Enterprise"),
            "address": lines(company.get("address")),
            "email": text(company.get("email")),
            "phone": text(company.get("phone")),
            "gstin": text(company.get("gstin")),
            "iec": text(company.get("iec")),
            "logo_path": logo_path,
            "signature_path": signature_path,
        },
        "invoice": {
            "title": text(invoice.get("title"), "Export Invoice"),
            "subtitle": text(invoice.get("subtitle"), "Supply under Letter of Undertaking without payment of Integrated Tax (IGST)"),
            "number": text(invoice.get("number")),
            "date": text(invoice.get("date")),
            "lut_no": text(invoice.get("lut_no")),
            "country": text(invoice.get("country"), "India"),
            "state_code": text(invoice.get("state_code")),
            "port_loading": text(invoice.get("port_loading")),
            "port_discharge": text(invoice.get("port_discharge")),
            "final_destination": text(invoice.get("final_destination")),
            "incoterm": text(invoice.get("incoterm")),
            "payment_terms": text(invoice.get("payment_terms")),
            "packages": text(invoice.get("packages")),
            "currency": text(invoice.get("currency"), "USD"),
            "declaration": text(invoice.get("declaration"), "Certified that the particulars given above are true and correct."),
            "amount_words": amount_words,
            "items": invoice_items,
        },
        "consignee": {
            "name": text(consignee.get("name")),
            "address": lines(consignee.get("address")),
            "registration": text(consignee.get("registration")),
            "attention": text(consignee.get("attention")),
            "phone": text(consignee.get("phone")),
        },
        "notify": {
            "name": text(notify.get("name")),
            "address": lines(notify.get("address")),
            "registration": text(notify.get("registration")),
            "attention": text(notify.get("attention")),
            "phone": text(notify.get("phone")),
        },
        "bank": {
            "account_name": text(bank.get("account_name")),
            "bank_name": text(bank.get("bank_name")),
            "account_number": text(bank.get("account_number")),
            "ifsc": text(bank.get("ifsc")),
            "swift": text(bank.get("swift")),
            "address": lines(bank.get("address")),
        },
        "packing_list": {
            "title": text(packing.get("title"), "Packing List"),
            "invoice_date": text(packing.get("invoice_date")) or text(invoice.get("date")),
            "invoice_number": text(packing.get("invoice_number")) or text(invoice.get("number")),
            "other_reference": text(packing.get("other_reference")) or f"IEC: {text(company.get('iec'))}",
            "buyer_name": text(packing.get("buyer_name")),
            "buyer_lines": lines(packing.get("buyer_lines")),
            "origin_state_country": text(packing.get("origin_state_country")) or text(invoice.get("port_loading")),
            "final_destination": text(packing.get("final_destination")) or text(invoice.get("final_destination")),
            "terms_lines": lines(packing.get("terms_lines")) or [f"{text(invoice.get('incoterm'))} {text(invoice.get('port_discharge')).upper()}".strip(), text(invoice.get("payment_terms"))],
            "pre_carriage_by": text(packing.get("pre_carriage_by"), "By Road"),
            "place_of_receipt": text(packing.get("place_of_receipt")),
            "vessel_flight_no": text(packing.get("vessel_flight_no"), "-"),
            "port_loading": text(packing.get("port_loading")) or text(invoice.get("port_loading")),
            "port_discharge": text(packing.get("port_discharge")) or text(invoice.get("port_discharge")),
            "rows": normalized_packing_rows,
            "case_no": text(packing.get("case_no")),
            "no_of_pcs": text(packing.get("no_of_pcs")),
            "total_quantity_mt": to_decimal(packing.get("total_quantity_mt"), "0"),
            "total_net_weight_ton": to_decimal(packing.get("total_net_weight_ton"), "0"),
            "total_gross_weight_ton": to_decimal(packing.get("total_gross_weight_ton"), "0"),
            "declaration": text(
                packing.get("declaration"),
                "We declare that this invoice shows the actual price of goods described and all particulars are true & correct.",
            ),
        },
    }


def build_company_header_lines(company: dict) -> list[str]:
    return [company["name"].upper(), *company["address"]]


def create_invoice_sheet(ws, ctx: dict) -> None:
    company = ctx["company"]
    invoice = ctx["invoice"]
    consignee = ctx["consignee"]
    notify = ctx["notify"]
    bank = ctx["bank"]
    ws.sheet_properties.tabColor = "10263A"
    widths = [4.5, 10.5, 10.5, 10.5, 8.5, 8.5, 8.5, 8.0, 8.0, 8.5]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    for row in range(1, 37):
        ws.row_dimensions[row].height = 18
    row_heights = {
        1: 16,
        2: 18,
        3: 22,
        4: 24,
        5: 18,
        6: 18,
        8: 18,
        9: 18,
        10: 18,
        11: 18,
        12: 18,
        14: 18,
        15: 18,
        16: 18,
        17: 18,
        18: 18,
        21: 18,
        23: 22,
        24: 46,
        25: 24,
        26: 24,
        27: 24,
        29: 18,
        30: 28,
        31: 18,
        32: 18,
        33: 18,
        34: 18,
        35: 18,
        36: 22,
    }
    for row, height in row_heights.items():
        ws.row_dimensions[row].height = height

    # Header
    apply_premium_range(ws, "A1:J6", fill=NAVY_FILL, alignment=CENTER, outer_side=DARK, inner_side=DARK)
    merge_premium(ws, "B2:E2", "COMMERCIAL EXPORT DOCUMENT", fill=NAVY_FILL, font=PREMIUM_META_FONT, alignment=LEFT_TOP, outer_side=DARK, inner_side=DARK)
    merge_premium(ws, "B3:G4", invoice["title"], fill=NAVY_FILL, font=PREMIUM_TITLE_FONT, alignment=LEFT_TOP, outer_side=DARK, inner_side=DARK)
    merge_premium(ws, "B5:G5", invoice["subtitle"], fill=NAVY_FILL, font=PREMIUM_MUTED_FONT, alignment=LEFT_TOP, outer_side=DARK, inner_side=DARK)
    location_line = ", ".join(company["address"][1:]) if len(company["address"]) > 1 else company["name"]
    merge_premium(ws, "B6:G6", location_line, fill=NAVY_FILL, font=PREMIUM_MUTED_FONT, alignment=LEFT_TOP, outer_side=DARK, inner_side=DARK)
    merge_premium(ws, "H2:J4", "", fill=PANEL_FILL, font=PREMIUM_BODY_FONT, alignment=CENTER, outer_side=SOFT, inner_side=SOFT)
    merge_premium(ws, "H5:J5", short_brand_name(company["name"]), fill=PANEL_FILL, font=PREMIUM_BODY_BOLD, alignment=CENTER, outer_side=SOFT, inner_side=SOFT)
    merge_premium(ws, "H6:J6", company["address"][-1] if company["address"] else "India", fill=NAVY_FILL, font=PREMIUM_MUTED_FONT, alignment=CENTER, outer_side=DARK, inner_side=DARK)
    add_image(ws, company["logo_path"], "H2", 86, 40)

    # Summary cards
    apply_premium_range(ws, "A8:C12", fill=PANEL_FILL, alignment=LEFT_TOP)
    merge_premium(ws, "A8:C8", "DOCUMENT SUMMARY", fill=PANEL_FILL, font=PREMIUM_CARD_TITLE_FONT, alignment=LEFT_TOP)
    merge_premium(ws, "A9:A9", "Invoice No", fill=PANEL_FILL, font=PREMIUM_SMALL_FONT, alignment=LEFT_TOP)
    merge_premium(ws, "B9:C9", invoice["number"], fill=PREMIUM_EDIT_FILL, font=PREMIUM_BODY_BOLD, alignment=LEFT_TOP)
    merge_premium(ws, "A10:A10", "Invoice Date", fill=PANEL_FILL, font=PREMIUM_SMALL_FONT, alignment=LEFT_TOP)
    merge_premium(ws, "B10:C10", invoice["date"], fill=PREMIUM_EDIT_FILL, font=PREMIUM_BODY_FONT, alignment=LEFT_TOP)
    merge_premium(ws, "A11:A11", "LUT No", fill=PANEL_FILL, font=PREMIUM_SMALL_FONT, alignment=LEFT_TOP)
    merge_premium(ws, "B11:C11", invoice["lut_no"], fill=PREMIUM_EDIT_FILL, font=PREMIUM_BODY_FONT, alignment=LEFT_TOP)
    merge_premium(ws, "A12:C12", "", fill=PANEL_FILL, font=PREMIUM_BODY_FONT, alignment=LEFT_TOP)

    apply_premium_range(ws, "D8:F12", fill=ACCENT_FILL, alignment=LEFT_TOP)
    merge_premium(ws, "D8:F8", "SHIPMENT ROUTE", fill=ACCENT_FILL, font=PREMIUM_CARD_TITLE_FONT, alignment=LEFT_TOP)
    merge_premium(ws, "D9:D9", "Port of Loading", fill=ACCENT_FILL, font=PREMIUM_SMALL_FONT, alignment=LEFT_TOP)
    merge_premium(ws, "E9:F9", invoice["port_loading"], fill=PREMIUM_EDIT_FILL, font=PREMIUM_BODY_FONT, alignment=LEFT_TOP)
    merge_premium(ws, "D10:D10", "Port of Discharge", fill=ACCENT_FILL, font=PREMIUM_SMALL_FONT, alignment=LEFT_TOP)
    merge_premium(ws, "E10:F10", invoice["port_discharge"], fill=PREMIUM_EDIT_FILL, font=PREMIUM_BODY_FONT, alignment=LEFT_TOP)
    merge_premium(ws, "D11:D11", "Final Destination", fill=ACCENT_FILL, font=PREMIUM_SMALL_FONT, alignment=LEFT_TOP)
    merge_premium(ws, "E11:F11", invoice["final_destination"], fill=PREMIUM_EDIT_FILL, font=PREMIUM_BODY_FONT, alignment=LEFT_TOP)
    merge_premium(ws, "D12:F12", "", fill=ACCENT_FILL, font=PREMIUM_BODY_FONT, alignment=LEFT_TOP)

    apply_premium_range(ws, "G8:J12", fill=PANEL_ALT_FILL, alignment=LEFT_TOP)
    merge_premium(ws, "G8:J8", "COMPLIANCE", fill=PANEL_ALT_FILL, font=PREMIUM_CARD_TITLE_FONT, alignment=LEFT_TOP)
    merge_premium(ws, "G9:H9", "GSTIN", fill=PANEL_ALT_FILL, font=PREMIUM_SMALL_FONT, alignment=LEFT_TOP)
    merge_premium(ws, "I9:J9", company.get("gstin", ""), fill=PREMIUM_EDIT_FILL, font=PREMIUM_BODY_FONT, alignment=LEFT_TOP)
    merge_premium(ws, "G10:H10", "IEC", fill=PANEL_ALT_FILL, font=PREMIUM_SMALL_FONT, alignment=LEFT_TOP)
    merge_premium(ws, "I10:J10", company.get("iec", ""), fill=PREMIUM_EDIT_FILL, font=PREMIUM_BODY_FONT, alignment=LEFT_TOP)
    merge_premium(ws, "G11:H11", "Origin / State / Code", fill=PANEL_ALT_FILL, font=PREMIUM_SMALL_FONT, alignment=LEFT_TOP)
    merge_premium(ws, "I11:J11", " / ".join(part for part in [invoice["country"], invoice["state_code"]] if part), fill=PREMIUM_EDIT_FILL, font=PREMIUM_BODY_FONT, alignment=LEFT_TOP)
    merge_premium(ws, "G12:J12", "", fill=PANEL_ALT_FILL, font=PREMIUM_BODY_FONT, alignment=LEFT_TOP)

    # Party cards
    apply_premium_range(ws, "A14:E18", fill=WHITE_FILL, alignment=LEFT_TOP)
    merge_premium(ws, "A14:E14", "CONSIGNEE", fill=WHITE_FILL, font=PREMIUM_CARD_TITLE_FONT, alignment=LEFT_TOP)
    consignee_text = "\n".join(
        part
        for part in [consignee["name"], *consignee["address"], consignee["registration"], consignee["attention"], consignee["phone"]]
        if part
    )
    merge_premium(ws, "A15:E18", consignee_text, fill=PREMIUM_EDIT_FILL, font=PREMIUM_BODY_FONT, alignment=LEFT_TOP)

    apply_premium_range(ws, "F14:J18", fill=WHITE_FILL, alignment=LEFT_TOP)
    merge_premium(ws, "F14:J14", "NOTIFY", fill=WHITE_FILL, font=PREMIUM_CARD_TITLE_FONT, alignment=LEFT_TOP)
    notify_text = "\n".join(
        part
        for part in [notify["name"], *notify["address"], notify["registration"], notify["attention"], notify["phone"]]
        if part
    )
    merge_premium(ws, "F15:J18", notify_text, fill=PREMIUM_EDIT_FILL, font=PREMIUM_BODY_FONT, alignment=LEFT_TOP)

    # Commercial strip
    merge_premium(ws, "A20:C20", "INCOTERM", fill=GOLD_FILL, font=PREMIUM_CARD_TITLE_FONT, alignment=LEFT_TOP)
    merge_premium(ws, "A21:C21", invoice["incoterm"], fill=GOLD_FILL, font=PREMIUM_BODY_BOLD, alignment=LEFT_TOP)
    merge_premium(ws, "D20:F20", "PAYMENT TERMS", fill=PANEL_ALT_FILL, font=PREMIUM_CARD_TITLE_FONT, alignment=LEFT_TOP)
    merge_premium(ws, "D21:F21", invoice["payment_terms"], fill=PANEL_ALT_FILL, font=PREMIUM_BODY_BOLD, alignment=LEFT_TOP)
    merge_premium(ws, "G20:J20", "PACKAGES", fill=ACCENT_FILL, font=PREMIUM_CARD_TITLE_FONT, alignment=LEFT_TOP)
    merge_premium(ws, "G21:J21", invoice["packages"], fill=ACCENT_FILL, font=PREMIUM_BODY_BOLD, alignment=LEFT_TOP)

    # Items table
    item_headers = [
        ("A23:A23", "Sr.No"),
        ("B23:D23", "Item Description"),
        ("E23:E23", "Quantity\n(Metric Ton)"),
        ("F23:F23", "Rate Per\nTon"),
        ("G23:G23", "Amount"),
        ("H23:H23", "IGST"),
        ("I23:J23", "Total USD"),
    ]
    for cell_range, label in item_headers:
        merge_premium(ws, cell_range, label, fill=NAVY_FILL, font=PREMIUM_TABLE_HEADER_FONT, alignment=CENTER, outer_side=DARK, inner_side=DARK)

    item_start = 24
    max_rows = 4
    for idx in range(max_rows):
        row = item_start + idx
        merge_premium(ws, f"A{row}:A{row}", idx + 1 if idx < len(invoice["items"]) else "", fill=PREMIUM_EDIT_FILL, font=PREMIUM_BODY_FONT, alignment=CENTER)
        merge_premium(ws, f"B{row}:D{row}", "", fill=PREMIUM_EDIT_FILL, font=PREMIUM_BODY_FONT, alignment=LEFT_TOP)
        merge_premium(ws, f"E{row}:E{row}", "", fill=PREMIUM_EDIT_FILL, font=PREMIUM_BODY_FONT, alignment=CENTER)
        merge_premium(ws, f"F{row}:F{row}", "", fill=PREMIUM_EDIT_FILL, font=PREMIUM_BODY_FONT, alignment=CENTER)
        merge_premium(ws, f"G{row}:G{row}", f'=IF(OR(E{row}="",F{row}=""),"",E{row}*F{row})', fill=PREMIUM_FORMULA_FILL, font=PREMIUM_BODY_FONT, alignment=RIGHT_CENTER)
        merge_premium(ws, f"H{row}:H{row}", "", fill=PREMIUM_EDIT_FILL, font=PREMIUM_BODY_FONT, alignment=CENTER)
        merge_premium(ws, f"I{row}:J{row}", f'=IF(G{row}="","",G{row}*(1+IF(H{row}="",0,H{row}/100)))', fill=PREMIUM_FORMULA_FILL, font=PREMIUM_BODY_FONT, alignment=RIGHT_CENTER)
        ws[f"E{row}"].number_format = '0.###'
        ws[f"F{row}"].number_format = '#,##0.00'
        ws[f"G{row}"].number_format = '#,##0.00'
        ws[f"H{row}"].number_format = '0.##'
        ws[f"I{row}"].number_format = '#,##0.00'

    for idx, item in enumerate(invoice["items"][:max_rows]):
        row = item_start + idx
        description = "\n".join([item["product"], *item["description_lines"]])
        ws[f"A{row}"] = item["index"]
        ws[f"B{row}"] = description
        ws[f"E{row}"] = float(item["quantity"])
        ws[f"F{row}"] = float(item["rate"])
        ws[f"H{row}"] = float(item["igst_rate"])

    # Bottom cards
    merge_premium(ws, "A29:E29", "TOTAL IN WORDS", fill=GOLD_FILL, font=PREMIUM_CARD_TITLE_FONT, alignment=LEFT_TOP)
    merge_premium(ws, "A30:E30", invoice["amount_words"], fill=GOLD_FILL, font=PREMIUM_BODY_BOLD, alignment=LEFT_TOP)

    merge_premium(ws, "F29:I29", "Total Quantity in Metric Tons", fill=WHITE_FILL, font=PREMIUM_CARD_TITLE_FONT, alignment=LEFT_TOP)
    merge_premium(ws, "J29:J29", "=SUM(E24:E27)", fill=PREMIUM_FORMULA_FILL, font=PREMIUM_TOTAL_FONT, alignment=RIGHT_CENTER)
    ws["J29"].number_format = '0.###'
    merge_premium(ws, "F30:I30", "Total Amount before Tax", fill=WHITE_FILL, font=PREMIUM_CARD_TITLE_FONT, alignment=LEFT_TOP)
    merge_premium(ws, "J30:J30", "=SUM(G24:G27)", fill=PREMIUM_FORMULA_FILL, font=PREMIUM_TOTAL_FONT, alignment=RIGHT_CENTER)
    ws["J30"].number_format = '#,##0.00'
    merge_premium(ws, "F31:I31", "Add: IGST", fill=WHITE_FILL, font=PREMIUM_CARD_TITLE_FONT, alignment=LEFT_TOP)
    merge_premium(ws, "J31:J31", "=SUMPRODUCT(G24:G27,H24:H27/100)", fill=PREMIUM_FORMULA_FILL, font=PREMIUM_TOTAL_FONT, alignment=RIGHT_CENTER)
    ws["J31"].number_format = '#,##0.00'
    merge_premium(ws, "F32:I32", "Total Amount after Tax", fill=WHITE_FILL, font=PREMIUM_CARD_TITLE_FONT, alignment=LEFT_TOP)
    merge_premium(ws, "J32:J32", "=SUM(I24:I27)", fill=PREMIUM_FORMULA_FILL, font=PREMIUM_TOTAL_FONT, alignment=RIGHT_CENTER)
    ws["J32"].number_format = '#,##0.00'

    merge_premium(ws, "A31:E31", "BANK DETAILS", fill=PANEL_ALT_FILL, font=PREMIUM_CARD_TITLE_FONT, alignment=LEFT_TOP)
    bank_pairs = [
        ("A32:B32", "Account Name", "C32:E32", bank["account_name"]),
        ("A33:B33", "Bank Name", "C33:E33", bank["bank_name"]),
        ("A34:B34", "Account Number", "C34:E34", bank["account_number"]),
        ("A35:B35", "IFSC / SWIFT", "C35:E35", " / ".join(part for part in [bank["ifsc"], bank["swift"]] if part)),
        ("A36:B36", "Bank Address", "C36:E36", ", ".join(bank["address"])),
    ]
    for label_range, label, value_range, value in bank_pairs:
        merge_premium(ws, label_range, label, fill=WHITE_FILL, font=PREMIUM_SMALL_FONT, alignment=LEFT_TOP)
        merge_premium(ws, value_range, value, fill=PREMIUM_EDIT_FILL, font=PREMIUM_BODY_FONT, alignment=LEFT_TOP)

    merge_premium(ws, "F33:J33", "DECLARATION", fill=PANEL_ALT_FILL, font=PREMIUM_CARD_TITLE_FONT, alignment=LEFT_TOP)
    declaration_text = f"{invoice['declaration']}\n\nFor {company['name']}\nAuthorized Signatory"
    merge_premium(ws, "F34:J36", declaration_text, fill=WHITE_FILL, font=PREMIUM_BODY_BOLD, alignment=LEFT_TOP)
    add_image(ws, company["signature_path"], "I34", 110, 42)

    configure_sheet(ws, "A1:J36")


def create_packing_list_sheet(ws, ctx: dict) -> None:
    company = ctx["company"]
    packing = ctx["packing_list"]
    consignee = ctx["consignee"]
    notify = ctx["notify"]
    ws.sheet_properties.tabColor = "2F6F67"
    widths = [5.5, 9.5, 9.5, 9.5, 9.5, 8.5, 8.5, 8.5, 8.0, 8.0]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    for row in range(1, 37):
        ws.row_dimensions[row].height = 18
    row_heights = {
        1: 16,
        2: 18,
        3: 22,
        4: 24,
        5: 18,
        6: 18,
        8: 18,
        9: 20,
        10: 20,
        11: 20,
        12: 20,
        14: 18,
        15: 18,
        16: 18,
        17: 18,
        18: 18,
        23: 22,
        24: 42,
        25: 24,
        26: 24,
        27: 24,
        31: 18,
        32: 28,
        33: 18,
        34: 18,
        35: 18,
        36: 24,
    }
    for row, height in row_heights.items():
        ws.row_dimensions[row].height = height

    # Header
    apply_premium_range(ws, "A1:J6", fill=NAVY_FILL, alignment=CENTER, outer_side=DARK, inner_side=DARK)
    merge_premium(ws, "B2:E2", "EXPORT SHIPPING DOCUMENT", fill=NAVY_FILL, font=PREMIUM_META_FONT, alignment=LEFT_TOP, outer_side=DARK, inner_side=DARK)
    merge_premium(ws, "B3:G4", packing["title"], fill=NAVY_FILL, font=PREMIUM_TITLE_FONT, alignment=LEFT_TOP, outer_side=DARK, inner_side=DARK)
    subtitle = "Editable packing list aligned for A4 print and shipment documentation"
    merge_premium(ws, "B5:G5", subtitle, fill=NAVY_FILL, font=PREMIUM_MUTED_FONT, alignment=LEFT_TOP, outer_side=DARK, inner_side=DARK)
    merge_premium(ws, "B6:G6", ", ".join(company["address"][1:]) if len(company["address"]) > 1 else company["name"], fill=NAVY_FILL, font=PREMIUM_MUTED_FONT, alignment=LEFT_TOP, outer_side=DARK, inner_side=DARK)
    merge_premium(ws, "H2:J4", "", fill=PANEL_FILL, font=PREMIUM_BODY_FONT, alignment=CENTER, outer_side=SOFT, inner_side=SOFT)
    merge_premium(ws, "H5:J5", short_brand_name(company["name"]), fill=PANEL_FILL, font=PREMIUM_BODY_BOLD, alignment=CENTER, outer_side=SOFT, inner_side=SOFT)
    merge_premium(ws, "H6:J6", company["address"][-1] if company["address"] else "India", fill=NAVY_FILL, font=PREMIUM_MUTED_FONT, alignment=CENTER, outer_side=DARK, inner_side=DARK)
    add_image(ws, company["logo_path"], "H2", 86, 40)

    exporter_lines = [f"M/S. {company['name']}", *company["address"]]
    if company["email"] or company["phone"]:
        exporter_lines.append(" | ".join(part for part in [company["email"], company["phone"]] if part))

    # Top cards
    apply_premium_range(ws, "A8:D12", fill=PANEL_FILL, alignment=LEFT_TOP)
    merge_premium(ws, "A8:D8", "EXPORTER", fill=PANEL_FILL, font=PREMIUM_CARD_TITLE_FONT, alignment=LEFT_TOP)
    merge_premium(ws, "A9:D12", "\n".join(exporter_lines), fill=PREMIUM_EDIT_FILL, font=PREMIUM_BODY_FONT, alignment=LEFT_TOP)

    apply_premium_range(ws, "E8:G12", fill=PANEL_ALT_FILL, alignment=LEFT_TOP)
    merge_premium(ws, "E8:G8", "DOCUMENT REFERENCE", fill=PANEL_ALT_FILL, font=PREMIUM_CARD_TITLE_FONT, alignment=LEFT_TOP)
    merge_premium(ws, "E9:F9", "Invoice Date", fill=PANEL_ALT_FILL, font=PREMIUM_SMALL_FONT, alignment=LEFT_TOP)
    merge_premium(ws, "G9:G9", packing["invoice_date"], fill=PREMIUM_EDIT_FILL, font=PREMIUM_BODY_FONT, alignment=LEFT_TOP)
    merge_premium(ws, "E10:F10", "Invoice No", fill=PANEL_ALT_FILL, font=PREMIUM_SMALL_FONT, alignment=LEFT_TOP)
    merge_premium(ws, "G10:G10", packing["invoice_number"], fill=PREMIUM_EDIT_FILL, font=PREMIUM_BODY_BOLD, alignment=LEFT_TOP)
    merge_premium(ws, "E11:F11", "Other Reference", fill=PANEL_ALT_FILL, font=PREMIUM_SMALL_FONT, alignment=LEFT_TOP)
    merge_premium(ws, "G11:G11", packing["other_reference"], fill=PREMIUM_EDIT_FILL, font=PREMIUM_BODY_FONT, alignment=LEFT_TOP)
    buyer_text = "\n".join(part for part in [packing["buyer_name"], *packing["buyer_lines"]] if part) or "Buyer same as consignee unless specified."
    merge_premium(ws, "E12:G12", buyer_text, fill=PANEL_ALT_FILL, font=PREMIUM_SMALL_FONT, alignment=LEFT_TOP)

    apply_premium_range(ws, "H8:J12", fill=ACCENT_FILL, alignment=LEFT_TOP)
    merge_premium(ws, "H8:J8", "SHIPMENT ROUTE", fill=ACCENT_FILL, font=PREMIUM_CARD_TITLE_FONT, alignment=LEFT_TOP)
    route_text = "\n".join(
        [
            f"Origin: {packing['origin_state_country']}",
            f"Final Destination: {packing['final_destination']}",
            f"Port Loading: {packing['port_loading']}",
            f"Port Discharge: {packing['port_discharge']}",
        ]
    )
    merge_premium(ws, "H9:J12", route_text, fill=PREMIUM_EDIT_FILL, font=PREMIUM_BODY_FONT, alignment=LEFT_TOP)

    # Parties
    apply_premium_range(ws, "A14:E18", fill=WHITE_FILL, alignment=LEFT_TOP)
    merge_premium(ws, "A14:E14", "CONSIGNEE", fill=WHITE_FILL, font=PREMIUM_CARD_TITLE_FONT, alignment=LEFT_TOP)
    consignee_text = "\n".join(
        part
        for part in [consignee["name"], *consignee["address"], consignee["registration"], consignee["attention"], consignee["phone"]]
        if part
    )
    merge_premium(ws, "A15:E18", consignee_text, fill=PREMIUM_EDIT_FILL, font=PREMIUM_BODY_FONT, alignment=LEFT_TOP)

    apply_premium_range(ws, "F14:J18", fill=WHITE_FILL, alignment=LEFT_TOP)
    merge_premium(ws, "F14:J14", "NOTIFY", fill=WHITE_FILL, font=PREMIUM_CARD_TITLE_FONT, alignment=LEFT_TOP)
    notify_text = "\n".join(
        part
        for part in [notify["name"], *notify["address"], notify["registration"], notify["attention"], notify["phone"]]
        if part
    )
    merge_premium(ws, "F15:J18", notify_text, fill=PREMIUM_EDIT_FILL, font=PREMIUM_BODY_FONT, alignment=LEFT_TOP)

    # Commercial strip
    merge_premium(ws, "A20:C20", "PRE-CARRIAGE BY", fill=GOLD_FILL, font=PREMIUM_CARD_TITLE_FONT, alignment=LEFT_TOP)
    merge_premium(ws, "A21:C21", packing["pre_carriage_by"], fill=GOLD_FILL, font=PREMIUM_BODY_BOLD, alignment=LEFT_TOP)
    merge_premium(ws, "D20:F20", "PACKAGES", fill=PANEL_ALT_FILL, font=PREMIUM_CARD_TITLE_FONT, alignment=LEFT_TOP)
    merge_premium(ws, "D21:F21", ctx["invoice"]["packages"], fill=PANEL_ALT_FILL, font=PREMIUM_BODY_BOLD, alignment=LEFT_TOP)
    merge_premium(ws, "G20:J20", "TERMS OF DELIVERY & PAYMENT", fill=ACCENT_FILL, font=PREMIUM_CARD_TITLE_FONT, alignment=LEFT_TOP)
    merge_premium(ws, "G21:J21", "\n".join(part for part in packing["terms_lines"] if part), fill=ACCENT_FILL, font=PREMIUM_BODY_BOLD, alignment=LEFT_TOP)

    # Goods table
    headers = [
        ("A23:B23", "Marks & Nos.\nContainer No."),
        ("C23:D23", "No. & Kind of\nPackages"),
        ("E23:H23", "Description of Goods"),
        ("I23:J23", "Quantity MT\n(Nt. Weight)"),
    ]
    for cell_range, label in headers:
        merge_premium(ws, cell_range, label, fill=NAVY_FILL, font=PREMIUM_TABLE_HEADER_FONT, alignment=CENTER, outer_side=DARK, inner_side=DARK)

    packing_start = 24
    max_rows = 4
    for idx in range(max_rows):
        row = packing_start + idx
        merge_premium(ws, f"A{row}:B{row}", "", fill=PREMIUM_EDIT_FILL, font=PREMIUM_BODY_FONT, alignment=LEFT_TOP)
        merge_premium(ws, f"C{row}:D{row}", "", fill=PREMIUM_EDIT_FILL, font=PREMIUM_BODY_FONT, alignment=CENTER_TOP)
        merge_premium(ws, f"E{row}:H{row}", "", fill=PREMIUM_EDIT_FILL, font=PREMIUM_BODY_FONT, alignment=LEFT_TOP)
        merge_premium(ws, f"I{row}:J{row}", "", fill=PREMIUM_EDIT_FILL, font=PREMIUM_BODY_FONT, alignment=CENTER_TOP)
        ws[f"I{row}"].number_format = '0.###'

    for idx, row_data in enumerate(packing["rows"][:max_rows]):
        row = packing_start + idx
        ws[f"A{row}"] = row_data["marks_numbers"]
        ws[f"C{row}"] = row_data["packages"]
        ws[f"E{row}"] = "\n".join(row_data["description_lines"])
        ws[f"I{row}"] = float(row_data["quantity_mt"])

    # Totals strip
    merge_premium(ws, "A29:B29", "Case No", fill=WHITE_FILL, font=PREMIUM_CARD_TITLE_FONT, alignment=LEFT_TOP)
    merge_premium(ws, "A30:B30", packing["case_no"], fill=PREMIUM_EDIT_FILL, font=PREMIUM_BODY_FONT, alignment=LEFT_TOP)
    merge_premium(ws, "C29:D29", "No. of Pcs", fill=WHITE_FILL, font=PREMIUM_CARD_TITLE_FONT, alignment=LEFT_TOP)
    merge_premium(ws, "C30:D30", packing["no_of_pcs"], fill=PREMIUM_EDIT_FILL, font=PREMIUM_BODY_FONT, alignment=LEFT_TOP)
    merge_premium(ws, "E29:F29", "Total Quantity (MT)", fill=WHITE_FILL, font=PREMIUM_CARD_TITLE_FONT, alignment=LEFT_TOP)
    merge_premium(ws, "E30:F30", "=SUM(I24:I27)", fill=PREMIUM_FORMULA_FILL, font=PREMIUM_TOTAL_FONT, alignment=CENTER)
    ws["E30"].number_format = '0.###'
    merge_premium(ws, "G29:H29", "Total Net Wt (TON)", fill=WHITE_FILL, font=PREMIUM_CARD_TITLE_FONT, alignment=LEFT_TOP)
    merge_premium(ws, "G30:H30", float(packing["total_net_weight_ton"]) if packing["total_net_weight_ton"] else "", fill=PREMIUM_EDIT_FILL, font=PREMIUM_TOTAL_FONT, alignment=CENTER)
    if packing["total_net_weight_ton"]:
        ws["G30"].number_format = '0.###'
    merge_premium(ws, "I29:J29", "Total Gross Wt (TON)", fill=WHITE_FILL, font=PREMIUM_CARD_TITLE_FONT, alignment=LEFT_TOP)
    merge_premium(ws, "I30:J30", float(packing["total_gross_weight_ton"]) if packing["total_gross_weight_ton"] else "", fill=PREMIUM_EDIT_FILL, font=PREMIUM_TOTAL_FONT, alignment=CENTER)
    if packing["total_gross_weight_ton"]:
        ws["I30"].number_format = '0.###'

    merge_premium(ws, "A32:F32", "DECLARATION", fill=PANEL_ALT_FILL, font=PREMIUM_CARD_TITLE_FONT, alignment=LEFT_TOP)
    merge_premium(ws, "A33:F36", packing["declaration"], fill=WHITE_FILL, font=PREMIUM_BODY_FONT, alignment=LEFT_TOP)
    merge_premium(ws, "G32:J32", "SIGNATURE", fill=PANEL_ALT_FILL, font=PREMIUM_CARD_TITLE_FONT, alignment=LEFT_TOP)
    merge_premium(ws, "G33:J36", f"For {company['name']}\n\nAuthorized Signatory", fill=WHITE_FILL, font=PREMIUM_BODY_BOLD, alignment=LEFT_TOP)
    add_image(ws, company["signature_path"], "H34", 110, 42)

    configure_sheet(ws, "A1:J36")


def build_workbook(data_path: Path, output_path: Path) -> None:
    data = json.loads(data_path.read_text())
    ctx = normalize_data(data, data_path)

    wb = Workbook()
    wb.calculation = CalcProperties(calcMode="auto", fullCalcOnLoad=True, forceFullCalc=True)

    ws_invoice = wb.active
    ws_invoice.title = "Export Invoice"
    create_invoice_sheet(ws_invoice, ctx)

    ws_packing = wb.create_sheet("Packing List")
    create_packing_list_sheet(ws_packing, ctx)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate editable Excel workbook for export invoice and packing list.")
    parser.add_argument("data_file", help="JSON data file")
    parser.add_argument("--output", help="Output .xlsx path")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    data_path = Path(args.data_file).expanduser().resolve()
    output_path = Path(args.output).expanduser().resolve() if args.output else OUTPUT_DIR / f"{data_path.stem}.xlsx"
    build_workbook(data_path, output_path)
    print(f"Workbook: {output_path}")


if __name__ == "__main__":
    main()
